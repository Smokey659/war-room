"""Personal CRM — ingest pipelines.

Three intake paths, all idempotent:

  1. ingest_sap_export(path)      — pipeline export .xlsx (deals + accounts)
  2. ingest_meeting_notes(recs)   — meeting-notes JSON (activities + contacts + reminders)
  3. extract_notes_from_text(txt) — LLM extraction: raw notes doc → meeting-notes JSON
                                    (used by the .docx/.md upload path)
  4. seed_from_vault()            — day-1 seed from the knowledge base: transcripts
                                    become activities; deal hubs attach to accounts

Field-ownership rule: exports refresh system fields on every run; user-owned
fields (status, motion, next_step, user_notes) are never written by ingest
after the row exists.
"""

from __future__ import annotations

import json
import re
import zipfile
from datetime import date, datetime
from pathlib import Path

from config import ANTHROPIC_API_KEY, MODEL, VAULT_PATH

from .db import conn_ctx, find_or_create_account, init_crm_db, log_import, now

# ---------------------------------------------------------------------------
# 1. Pipeline export (.xlsx)
# ---------------------------------------------------------------------------

# Expected header → internal field. Matching is fuzzy (lowercase, strip dots) so
# minor export-format drift doesn't break the parser.
EXPORT_COLUMNS = {
    "account id": "sap_account_id",
    "account name": "account_name",
    "opp id": "opp_id",
    "description": "name",
    "phase": "phase",
    "days in phase": "days_in_phase",
    "drm category": "forecast_cat",
    "revenue/closing date": "close_date",
    "closing quarter": "close_quarter",
    "opp owner": "opp_owner",
    "account owner": "account_owner",
    "untouched days": "untouched_days",
    "uw pipe (1x)": "amount_1x",
    "uw pipe (mp)": "amount_mp",
}

# Product inference from the description line, first match wins.
PRODUCT_PATTERNS = [
    (r"successfactors|sf hris|sf@|hcm", "SuccessFactors"),
    (r"s/4|s4|cloud erp|erp", "S/4HANA / ERP"),
    (r"ariba|procurement", "Ariba"),
    (r"leanix", "LeanIX"),
    (r"signavio|btm", "Signavio / BTM"),
    (r"\bbdc\b|businessobjects|bobj", "BDC / Analytics"),
    (r"gwsap|grow", "GROW"),
    (r"fieldglass", "FieldGlass"),
    (r"concur|t&e", "Concur"),
    (r"\blms\b|learning", "Learning"),
    (r"analytics|sac\b", "SAC"),
    (r"docusign", "DocuSign resale"),
    (r"ai\b|joule", "Business AI"),
]


def _infer_product(description: str) -> str | None:
    d = (description or "").lower()
    for pat, label in PRODUCT_PATTERNS:
        if re.search(pat, d):
            return label
    return None


def _coerce_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s[:10] if s else None


def _coerce_num(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def ingest_sap_export(path: Path, filename: str | None = None) -> dict:
    """Parse a pipeline export workbook and upsert accounts + deals."""
    import openpyxl  # deferred: heavy import

    init_crm_db()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"added": 0, "updated": 0, "skipped": 0, "error": "empty sheet"}

    # header row = first row containing an opp-id-ish column
    header_idx = None
    colmap: dict[int, str] = {}
    for i, row in enumerate(rows[:5]):
        cand = {}
        for j, cell in enumerate(row):
            if cell is None:
                continue
            key = re.sub(r"[.\s]+", " ", str(cell).strip().lower()).replace("opp.", "opp")
            key = key.replace("opp ", "opp ")  # normalize
            for known, field in EXPORT_COLUMNS.items():
                if key == known or key.replace(".", "") == known.replace(".", ""):
                    cand[j] = field
        if "opp_id" in cand.values():
            header_idx, colmap = i, cand
            break
    if header_idx is None:
        return {"added": 0, "updated": 0, "skipped": 0, "error": "no recognizable header row"}

    added = updated = skipped = 0
    with conn_ctx() as conn:
        for row in rows[header_idx + 1:]:
            rec = {field: row[j] for j, field in colmap.items() if j < len(row)}
            opp_id = str(rec.get("opp_id") or "").strip()
            if not opp_id:
                skipped += 1
                continue
            account_name = str(rec.get("account_name") or "").strip() or "(unknown account)"
            sap_acct = str(rec.get("sap_account_id") or "").strip() or None
            acct_id = find_or_create_account(conn, account_name, "sap-export", sap_acct)

            desc = str(rec.get("name") or "").strip() or opp_id
            passive = 1 if opp_id.startswith("006") else 0
            renewal = 1 if "renewal" in desc.lower() else 0
            system_fields = dict(
                name=desc,
                phase=str(rec.get("phase") or "").strip() or None,
                days_in_phase=int(_coerce_num(rec.get("days_in_phase")) or 0),
                forecast_cat=str(rec.get("forecast_cat") or "").strip() or None,
                close_date=_coerce_date(rec.get("close_date")),
                close_quarter=str(rec.get("close_quarter") or "").strip() or None,
                amount_1x=_coerce_num(rec.get("amount_1x")),
                amount_mp=_coerce_num(rec.get("amount_mp")),
                opp_owner=str(rec.get("opp_owner") or "").strip() or None,
                untouched_days=int(_coerce_num(rec.get("untouched_days")) or 0),
                passive=passive,
                renewal=renewal,
            )

            existing = conn.execute("SELECT id FROM deals WHERE opp_id=?", (opp_id,)).fetchone()
            if existing:
                sets = ", ".join(f"{k}=?" for k in system_fields)
                conn.execute(
                    f"UPDATE deals SET {sets}, account_id=?, updated_at=? WHERE opp_id=?",
                    (*system_fields.values(), acct_id, now(), opp_id))
                updated += 1
            else:
                product = _infer_product(desc)
                motion = "passive" if passive else None
                conn.execute(
                    """INSERT INTO deals (account_id, opp_id, name, product, phase, days_in_phase,
                        forecast_cat, close_date, close_quarter, amount_1x, amount_mp, opp_owner,
                        untouched_days, passive, renewal, motion, source, created_at, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (acct_id, opp_id, desc, product, system_fields["phase"],
                     system_fields["days_in_phase"], system_fields["forecast_cat"],
                     system_fields["close_date"], system_fields["close_quarter"],
                     system_fields["amount_1x"], system_fields["amount_mp"],
                     system_fields["opp_owner"], system_fields["untouched_days"],
                     passive, renewal, motion, "sap-export", now(), now()))
                added += 1
        log_import(conn, "sap-export", filename or path.name, added, updated, skipped)
    return {"added": added, "updated": updated, "skipped": skipped}


# ---------------------------------------------------------------------------
# 2. Meeting-notes JSON (the collector contract)
# ---------------------------------------------------------------------------

def ingest_meeting_notes(records: list[dict], filename: str | None = None) -> dict:
    """Upsert activities + contacts (+ reminders from next_steps).

    Record shape (all keys optional except date/account/title):
      {date, account, deal_hint, type, title, attendees:[{name,title,org}],
       summary, next_steps:[...], source, source_ref}
    """
    init_crm_db()
    added = updated = skipped = 0
    with conn_ctx() as conn:
        for rec in records:
            acct_name = (rec.get("account") or "").strip()
            title = (rec.get("title") or "").strip()
            d = (rec.get("date") or "").strip()[:10]
            if not (acct_name and title and d):
                skipped += 1
                continue
            acct_id = find_or_create_account(conn, acct_name, "meeting-notes")

            # link to a deal on the same account if the hint matches, else any single active deal
            deal_id = None
            hint = (rec.get("deal_hint") or "").lower()
            deals = conn.execute(
                "SELECT id, name, product FROM deals WHERE account_id=? AND passive=0", (acct_id,)).fetchall()
            if hint:
                for dl in deals:
                    if hint in (dl["name"] or "").lower() or hint in (dl["product"] or "").lower():
                        deal_id = dl["id"]
                        break
            if deal_id is None and len(deals) == 1:
                deal_id = deals[0]["id"]

            existing = conn.execute(
                "SELECT id FROM activities WHERE account_id=? AND date=? AND title=?",
                (acct_id, d, title)).fetchone()
            payload = (
                rec.get("type") or "meeting",
                (rec.get("summary") or "").strip() or None,
                json.dumps(rec.get("next_steps") or []),
                rec.get("source") or "meeting-notes",
                rec.get("source_ref"),
            )
            if existing:
                conn.execute(
                    """UPDATE activities SET deal_id=?, type=?, summary=?, next_steps=?,
                       source=?, source_ref=? WHERE id=?""",
                    (deal_id, *payload, existing["id"]))
                updated += 1
            else:
                conn.execute(
                    """INSERT INTO activities (account_id, deal_id, date, type, title, summary,
                       next_steps, source, source_ref, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (acct_id, deal_id, d, payload[0], title, payload[1], payload[2],
                     payload[3], payload[4], now()))
                added += 1
                # reminders from next steps (only on first ingest of the activity)
                for step in (rec.get("next_steps") or [])[:10]:
                    conn.execute(
                        """INSERT INTO reminders (account_id, deal_id, kind, note, source, created_at)
                           VALUES (?,?,?,?,?,?)""",
                        (acct_id, deal_id, "follow-up", str(step)[:300], "meeting-notes", now()))

            # contacts
            for att in rec.get("attendees") or []:
                nm = (att.get("name") or "").strip()
                if not nm:
                    continue
                org = att.get("org") or "customer"
                row = conn.execute(
                    "SELECT id, title FROM contacts WHERE account_id=? AND name=?",
                    (acct_id, nm)).fetchone()
                if row:
                    conn.execute(
                        "UPDATE contacts SET title=COALESCE(?, title), last_touch=MAX(COALESCE(last_touch,''), ?), updated_at=? WHERE id=?",
                        (att.get("title"), d, now(), row["id"]))
                else:
                    conn.execute(
                        """INSERT INTO contacts (account_id, name, title, org, last_touch, source,
                           created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)""",
                        (acct_id, nm, att.get("title"), org, d, "meeting-notes", now(), now()))
        log_import(conn, "meeting-notes", filename, added, updated, skipped)
    return {"added": added, "updated": updated, "skipped": skipped}


# ---------------------------------------------------------------------------
# 3. LLM extraction — raw notes document → meeting-notes JSON
# ---------------------------------------------------------------------------

EXTRACT_SYSTEM = """You convert raw sales meeting-notes documents into structured JSON.
Output ONLY a JSON array (no prose, no code fences). Each element:
{
  "date": "YYYY-MM-DD",              // the meeting date found in the document
  "account": "Customer account name",// the CUSTOMER organization, not the vendor
  "deal_hint": "short product/deal hint if evident",
  "type": "meeting",
  "title": "short title of the meeting",
  "attendees": [{"name": "...", "title": "...", "org": "customer|internal|partner"}],
  "summary": "5-8 sentence factual summary: what happened, what changed, key numbers, risks",
  "next_steps": ["action item with owner if stated", ...],
  "source": "notes-doc"
}
One document usually yields ONE record. Use "internal" org for the vendor's own
staff (account executives, solution advisors). Dates must be ISO format."""


def docx_to_text(path: Path) -> str:
    """Extract plain text from a .docx without extra dependencies."""
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
    xml = xml.replace("</w:p>", "\n")
    text = re.sub(r"<[^>]+>", "", xml)
    text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"')
    lines = [l.strip() for l in text.split("\n")]
    return "\n".join(l for l in lines if l)


def extract_notes_from_text(text: str) -> list[dict]:
    """Claude call: raw notes text → meeting-notes JSON records."""
    if not ANTHROPIC_API_KEY:
        raise RuntimeError("ANTHROPIC_API_KEY not configured — cannot run extraction")
    from anthropic import Anthropic

    client = Anthropic(api_key=ANTHROPIC_API_KEY)
    resp = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        system=EXTRACT_SYSTEM,
        messages=[{"role": "user", "content": text[:80000]}],
    )
    raw = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()
    raw = re.sub(r"^```(json)?|```$", "", raw, flags=re.MULTILINE).strip()
    records = json.loads(raw)
    if isinstance(records, dict):
        records = [records]
    return records


# ---------------------------------------------------------------------------
# 4. Vault seed — transcripts → activities; deal hubs → account/hub links
# ---------------------------------------------------------------------------

_TRANSCRIPT_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})-(customer|sales|partner|internal)-(.+)\.md$")


def seed_from_vault() -> dict:
    """Deterministic pass over the knowledge base.

    - transcripts/<any>/YYYY-MM-DD-<type>-<slug>.md  → one activity each
    - deal-hub notes (project-*-deal-*.md)           → attach hub path to matching account
    No LLM calls; safe to re-run.
    """
    init_crm_db()
    added = updated = skipped = 0
    detail = []

    transcripts_dir = VAULT_PATH / "transcripts"
    hub_glob = list(VAULT_PATH.glob("projects/**/active-deals/*.md"))

    with conn_ctx() as conn:
        # --- transcripts → activities (customer transcripts only: their slug IS the
        # account; sales/internal transcripts are mentor sessions, not deal activity) ---
        if transcripts_dir.exists():
            for p in sorted(transcripts_dir.rglob("*.md")):
                m = _TRANSCRIPT_RE.match(p.name)
                if not m or m.group(2) != "customer":
                    skipped += 1
                    continue
                d, kind, slug = m.groups()
                # Slugs often carry trailing person/context tokens ("<account>-<person>").
                # Try progressively shorter prefixes against existing accounts before
                # creating a new one, so "acme-jane-doe" lands on "City of Acme".
                tokens = slug.replace("-", " ").title().split()
                acct_name = " ".join(tokens)
                from .db import _norm  # local import to keep module top clean
                existing_accts = conn.execute("SELECT id, name FROM accounts").fetchall()
                matched = None
                for k in range(len(tokens), 0, -1):
                    cand = _norm(" ".join(tokens[:k]))
                    if len(cand) < 4 and k > 1:
                        continue
                    for r in existing_accts:
                        rn = _norm(r["name"])
                        if cand and (rn.startswith(cand) or cand.startswith(rn) or cand in rn):
                            matched = r["id"]
                            break
                    if matched:
                        break
                if matched:
                    acct_id = matched
                else:
                    acct_id = find_or_create_account(conn, acct_name, "vault-parse")
                # first markdown heading as title
                title = None
                try:
                    for line in p.read_text(errors="ignore").splitlines():
                        if line.startswith("#"):
                            title = line.lstrip("# ").strip()
                            break
                except OSError:
                    pass
                title = title or p.stem
                existing = conn.execute(
                    "SELECT id FROM activities WHERE account_id=? AND date=? AND title=?",
                    (acct_id, d, title)).fetchone()
                if existing:
                    updated += 1
                    continue
                conn.execute(
                    """INSERT INTO activities (account_id, date, type, title, summary, next_steps,
                       source, source_ref, created_at) VALUES (?,?,?,?,?,?,?,?,?)""",
                    (acct_id, d, "meeting" if kind == "customer" else kind, title,
                     None, "[]", "vault-parse", str(p.relative_to(VAULT_PATH)), now()))
                added += 1

        # --- deal hubs → source_hub on deals of the matching account ---
        for p in hub_glob:
            try:
                head = p.read_text(errors="ignore")[:2000]
            except OSError:
                continue
            m = re.search(r"^#\s.*?Deal:\s*(.+)$", head, flags=re.MULTILINE)
            if not m:
                continue
            # strip parentheticals and trailing state suffixes like ", WA"
            acct_name = m.group(1).strip().split("(")[0].strip().rstrip(",")
            acct_name = re.sub(r",\s*[A-Z]{2}$", "", acct_name).strip()
            acct_id = find_or_create_account(conn, acct_name, "vault-parse")
            rel = str(p.relative_to(VAULT_PATH))
            n_upd = conn.execute(
                "UPDATE deals SET source_hub=? WHERE account_id=? AND source_hub IS NULL",
                (rel, acct_id)).rowcount
            if n_upd:
                detail.append(f"hub→{acct_name}")
                updated += n_upd

        log_import(conn, "vault-seed", None, added, updated, skipped, ", ".join(detail[:20]))
    return {"added": added, "updated": updated, "skipped": skipped}
